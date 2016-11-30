"""
    Group Name: OASong
    Member 1.59070098 Paroot Satjawanit
           2.59070156 Vasanchai Prakobkij
    Program Name: Cigarette Analyze
"""
from pygal.style import Style
from openpyxl import load_workbook
from pygal.style import DarkStyle
import pygal
def region():
    wb = load_workbook(filename = 'region.xlsx')
    sheet_ranges = wb['Sheet1']
    level = []
    allr = []
    ktm = []
    mid = []
    north = []
    esan = []
    south = []
    for i in range(2, 12):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cell_c = 'C'+str(i)
        cell_d = 'D'+str(i)
        cell_e = 'E'+str(i)
        cell_f = 'F'+str(i)
        cell_g = 'G'+str(i)
        level.append(sheet_ranges[cell_a].value)
        ktm.append(sheet_ranges[cell_c].value)
        mid.append(sheet_ranges[cell_d].value)
        north.append(sheet_ranges[cell_e].value)
        esan.append(sheet_ranges[cell_f].value)
        south.append(sheet_ranges[cell_g].value)
    line_chart = pygal.Dot(fill=True, interpolate='cubic', style=DarkStyle)
    line_chart.title = 'Static of Smoker sorted by Region & Education'
    line_chart.x_labels = map(str, level)
    line_chart.add('กรุงเทพฯ', ktm)
    line_chart.add('กลาง', mid)
    line_chart.add('เหนือ', north)
    line_chart.add('อิสาน', esan)
    line_chart.add('ใต้', south)
    line_chart.render_to_file('Graph_region.svg')
    age()

def age():
    wb = load_workbook(filename = 'ages.xlsx')
    sheet_ranges = wb['Sheet1']
    level = []
    never = []
    always = []
    sometimes = []
    for i in range(2, 8):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cell_c = 'C'+str(i)
        cell_d = 'D'+str(i)
        level.append(sheet_ranges[cell_a].value)
        never.append(sheet_ranges[cell_b].value)
        always.append(sheet_ranges[cell_c].value)
        sometimes.append(sheet_ranges[cell_d].value)
    line_chart = pygal.HorizontalBar(fill=True, interpolate='cubic', style=DarkStyle)
    line_chart.title = 'Static of Smoker sorted by Age&Sex'
    line_chart.x_labels = map(str, level)
    line_chart.add('ไม่เคยสูบ', never)
    line_chart.add('สูบประจำ', always)
    line_chart.add('สูบนานๆครั้ง', sometimes)
    line_chart.render_to_file('Graph_age_sex.svg')
    job()

def job():
    wb = load_workbook(filename = 'jobs.xlsx')
    sheet_ranges = wb['Sheet1']
    jobs = []
    all_r = []
    ktm = []
    mid = []
    north = []
    esan = []
    south = []
    for i in range(2, 11):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cell_c = 'C'+str(i)
        cell_d = 'D'+str(i)
        cell_e = 'E'+str(i)
        cell_f = 'F'+str(i)
        cell_g = 'G'+str(i)
        jobs.append(sheet_ranges[cell_a].value)
        ktm.append(sheet_ranges[cell_c].value)
        mid.append(sheet_ranges[cell_d].value)
        north.append(sheet_ranges[cell_e].value)
        esan.append(sheet_ranges[cell_f].value)
        south.append(sheet_ranges[cell_g].value)
    line_chart = pygal.Bar(fill=True, interpolate='cubic', style=DarkStyle)
    line_chart.title = 'Static of Smoker sorted by jobs'
    line_chart.x_labels = map(str, jobs)
    line_chart.add('กรุงเทพมหานคร', ktm)
    line_chart.add('กลาง', mid)
    line_chart.add('เหนือ', north)
    line_chart.add('อีสาน', esan)
    line_chart.add('ใต้', south)
    line_chart.render_to_file('Graph_jobs.svg')
    quit()

def quit():
    wb = load_workbook(filename = 'quit.xlsx')
    sheet_ranges = wb['Sheet1']
    level = []
    form_a = []
    form_b = []
    form_c = []
    form_d = []
    form_e = []
    for i in range(2, 8):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cell_c = 'C'+str(i)
        cell_d = 'D'+str(i)
        cell_e = 'E'+str(i)
        cell_f = 'F'+str(i)
        level.append(sheet_ranges[cell_a].value)
        form_a.append(sheet_ranges[cell_b].value)
        form_b.append(sheet_ranges[cell_c].value)
        form_c.append(sheet_ranges[cell_d].value)
        form_d.append(sheet_ranges[cell_e].value)
        form_e.append(sheet_ranges[cell_f].value)
    line_chart = pygal.HorizontalBar(fill=True, interpolate='cubic', style=DarkStyle)
    line_chart.title = 'Static of Smoker who Stop Smoking\nSorted by sex&age'
    line_chart.x_labels = map(str, level)
    line_chart.add('ครอบครัวขอร้อง', form_a)
    line_chart.add('ประหยัดเงิน', form_b)
    line_chart.add('การรณรงค์', form_c)
    line_chart.add('สถานที่สูบไม่อำนวย', form_d)
    line_chart.add('ป่วย', form_e)
    line_chart.render_to_file('Graph_quit.svg')
    smoke()

def smoke():
    gauge = pygal.SolidGauge(half_pie=True, inner_radius=0.70,
        style=DarkStyle(value_font_size=15))
    gauge.title = 'Toxic in Smoke'
    wb = load_workbook(filename = 'smoke.xlsx')
    sheet_ranges = wb['Sheet1']
    level = []
    num = []
    sumall = sheet_ranges['B16'].value
    percent_formatter = lambda x: '{:.10g}%'.format(x)
    gauge.value_formatter = percent_formatter
    for i in range(1, 16):
        cell_a = 'A'+str(i)
        cell_c = 'C'+str(i)
        level.append(sheet_ranges[cell_a].value)
        num.append(float(sheet_ranges[cell_c].value))
    for i in range(15):
        gauge.add(level[i], [{'value': num[i], 'max_value': 100}])
    gauge.render_to_file('Graph_smoke.svg')
    store()

def store():
    wb = load_workbook(filename = 'store.xlsx')
    sheet_ranges = wb['Sheet1']
    age = []
    datab = []
    datac = []
    datad = []
    datae = []
    dataf = []
    for i in range(2, 6):
        cella = 'A'+str(i)
        cellb = 'B'+str(i)
        cellc = 'C'+str(i)
        celld = 'D'+str(i)
        celle = 'E'+str(i)
        cellf = 'F'+str(i)
        age.append(sheet_ranges[cella].value)
        datab.append(sheet_ranges[cellb].value)
        datac.append(sheet_ranges[cellc].value)
        datad.append(sheet_ranges[celld].value)
        datae.append(sheet_ranges[celle].value)
        dataf.append(sheet_ranges[cellf].value)
    line_chart = pygal.HorizontalBar(fill=True, interpolate='cubic', style=DarkStyle)
    line_chart.title = 'Stactic of Store where sold Cigarette'
    line_chart.x_labels = map(str, age)
    line_chart.add('ร้านขายของชำ', datab)
    line_chart.add('ร้านสะดวกซื้อ', datac)
    line_chart.add('ร้านค้าส่ง', datad)
    line_chart.add('แผงลอย', datae)
    line_chart.add('ต่างประเทศ', dataf)
    line_chart.render_to_file('Graph_store.svg')
    idcard()

def idcard():
    wb = load_workbook(filename = 'id.xlsx')
    sheet_ranges = wb['Sheet1']
    region = []
    datab = []
    datac = []
    datad = []
    datae = []
    dataf = []
    for i in range(2, 4):
        cella = 'A'+str(i)
        cellb = 'B'+str(i)
        cellc = 'C'+str(i)
        celld = 'D'+str(i)
        celle = 'E'+str(i)
        cellf = 'F'+str(i)
        region.append(sheet_ranges[cella].value)
        datab.append(sheet_ranges[cellb].value)
        datac.append(sheet_ranges[cellc].value)
        datad.append(sheet_ranges[celld].value)
        datae.append(sheet_ranges[celle].value)
        dataf.append(sheet_ranges[cellf].value)
    line_chart = pygal.StackedBar(fill=True, interpolate='cubic', style=DarkStyle)
    line_chart.title = 'Stactic of Seller who ask for ID card before sell a Cigarette'
    line_chart.x_labels = map(str, region)
    line_chart.add('กรุงเทพฯ', datab)
    line_chart.add('กลาง(ไม่รวมกทม.)', datac)
    line_chart.add('เหนือ', datad)
    line_chart.add('ตะวันออกเฉียงเหนือ', datae)
    line_chart.add('ใต้', dataf)
    line_chart.render_to_file('Graph_askid.svg')

region()
